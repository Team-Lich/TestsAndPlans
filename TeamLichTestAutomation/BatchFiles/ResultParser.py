import untangle
import sys
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

sourceFile = sys.argv[1]
timestamp = sys.argv[2]

doc = untangle.parse(sourceFile + ".trx")

total = doc.TestRun.ResultSummary.Counters["total"]
passed = doc.TestRun.ResultSummary.Counters["passed"]

print("\n{0} tests ran, {1} passed\n".format(total, passed))

results = doc.TestRun.Results.UnitTestResult

wb = Workbook()
ws = wb.active
ws.title = "TestResults"

currentRow = 1

for result in results:
	lineFormat = "Test {0} ran for {1}. Result: {2}"
	testName = result["testName"]
	testDuration = result["duration"]
	testOutcome = result["outcome"]	
	
	ws["A{0}".format(currentRow)] = testName
	ws["B{0}".format(currentRow)] = testDuration
	ws["C{0}".format(currentRow)] = testOutcome
	
	if testOutcome == "Passed":
		ft = Font(color='00237700')
	else:
		ft = Font(color='00D50000')
		
	ws["A{0}".format(currentRow)].font = ft
	ws["B{0}".format(currentRow)].font = ft
	ws["C{0}".format(currentRow)].font = ft
	
	currentRow += 1
	
	print(lineFormat.format(testName, testDuration, testOutcome))
	
wb.save(sourceFile + ".xlsx")
